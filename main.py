from bs4 import BeautifulSoup
import requests
import openpyxl as xl
import lxml

wb = xl.load_workbook("WA_Members.xlsx")
sheet = wb['Sheet1']

# Create variables for Excel file
excel_row = 2
member_cell = 1
committee_cell = 2

# Sets up the website to scrape and isolates the table with Member and Committee info
response = requests.get("https://app.leg.wa.gov/Rosters/CommitteeMembersByMember/House")
house_rep_page = response.text
soup = BeautifulSoup(house_rep_page, "lxml")
table = soup.find('table', {'class': 'tablesaw'}).find('tbody')

# Create headers in Excel file
sheet.cell(1, member_cell).value = 'Member'
sheet.cell(1, committee_cell).value = 'Committees'


# Loops through table to find member and committee information to write to excel file
for row in table.find_all('tr'):
    cells = row.find_all('td')
    member_name = cells[0].find('a').text.strip()
    committee_name = cells[1].text.strip()
    sheet.cell(excel_row, member_cell).value = member_name
    sheet.cell(excel_row, committee_cell).value = committee_name
    excel_row += 1

wb.save('WA_Members.xlsx')
